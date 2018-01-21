from openpyxl import Workbook, load_workbook
import json
import re
wb = load_workbook("waterstress.xlsx")
ws = wb.active
print("Sheet loaded")
def cv(row, col):
	col = str(col)
	if not col.isdigit():
		return ws[col + str(row)].value
	else:
		return ws.cell(int(row), int(col)).value
def cvn(row, col):
	retval = cv(row,col)
	if retval is None:
		#print('bad')
		return 0
	else:
		return int(retval)
#data format: {Code: {"name": Item Name, "countries":{China, 1050, United States:, 2080}}}
#eg 15[countries][China] = 1050
stress_data = {}
for r in range(2,169):
	stress_data[cv(r, 'B')] = cv(r, 'C')
#print(cvn(r,'J'))
print("Data Parsed")
with open('stress.json', 'w') as f:
	json.dump(stress_data, f)
