from openpyxl import Workbook, load_workbook
import json
import re
wb = load_workbook("impactByProduct.xlsx")
ws = wb.active
print("Sheet loaded")
def cv(row, col):
	if not col.isdigit():
		return ws[str(col) + str(row)].value
	else:
		return ws.cell(int(row), int(col)).value
def cvn(row, col):
	retval = cv(row,col)
	if retval is None:
		return 0
	else:
		return retval
water_data = {}
#re.split('\W+', cv(row, 'E'))[0]
def get_footprint(row, col):
	return {"Name": cv(row, 'E'), "Green": cvn(row,col), "Blue": cvn(row+1, col), "Grey": cvn(row+2, col)}
def get_region_data(column):
	data = {}
	for row in range(7, 1066):
		if cv(row, 'E') is not None: #only interested in broader FAOSTAT categories
			if not cv(row, column):
				continue
			data[cv(row, 'A')] = get_footprint(row, column)
	return data
water_data["Global"] = get_region_data("J")
print("Data parsed")
#print(cv(7, "DOF"))
with open('GlobalImpact.json', 'w') as f:
	json.dump(water_data, f)