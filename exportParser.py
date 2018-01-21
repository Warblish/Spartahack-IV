from openpyxl import Workbook, load_workbook
import json
import re
wb = load_workbook("FAOExport.xlsx")
ws = wb.active
print("Sheet loaded")
def cv(row, col):
	col = str(col)
	if not col.isdigit():
		return ws[col + str(row)].value
	else:
		return ws.cell(int(row), int(col)).value
def cvn(row, col):
	retval = cv(row,col)
	if retval is None:
		#print('bad')
		return 0
	else:
		return int(retval)
#data format: {Code: {"name": Item Name, "countries":{China, 1050, United States:, 2080}}}
#eg 15[countries][China] = 1050
export_data = {}
r = 2
#print(cvn(r,'J'))
current_code = 15
temp_data = {}
world_qty = 0
while r <= 94638:
	old_code = current_code
	current_code = cv(r, 'F')
	if current_code != old_code:
		export_data[old_code] = {"name": cv(r-1, 'G'), 
		"countries": {country: str(temp_data[country]*1.0/world_qty * 100) for country in temp_data if temp_data[country] and world_qty}}
		# for country in temp_data:
		# 	if temp_data[country] and world_qty:
		# 		print(temp_data[country]*1.0/world_qty)
		# print({"name": cv(r-1, 'G'), 
		# "countries": {country: temp_data[country]/world_qty for country in temp_data if temp_data[country] and world_qty}})
		# print(world_qty)
		temp_data = {}
		world_qty = 0
	if cv(r, 'E') == "Export Quantity": #only interested in exports
		temp_data[cv(r, 'C')] = cvn(r, 'J')
		world_qty += cvn(r, 'J')
		#print(temp_data)
	r += 1
export_data[current_code] = {"name": cv(r, 'G'), 
	"countries": {country: str(temp_data[country]*1.0/world_qty * 100) for country in temp_data if temp_data[country] and world_qty}}
print("Data Parsed")
with open('proportions.json', 'w') as f:
	json.dump(export_data, f)
